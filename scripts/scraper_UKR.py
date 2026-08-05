"""
scraper_UKR.py  –  Ukrán Oryx adatok scraper (javított verzió)

Használat:
    python scraper_UKR.py                    # default: data/xyz_UKR.xlsx
    python scraper_UKR.py data/xyz_UKR.xlsx  # explicit elérési út
"""

import sys, os, re, logging
import pandas as pd
import requests
from bs4 import BeautifulSoup

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from fetch_with_backoff import fetch  # noqa: E402
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)

URL = "https://www.oryxspioenkop.com/2022/02/attack-on-europe-documenting-ukrainian.html"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_FILE = os.path.normpath(os.path.join(SCRIPT_DIR, '..', 'data', 'xyz_UKR.xlsx'))

CATEGORIES = [
    ("Tanks", "T-55"),
    ("Armoured Fighting Vehicles", "BTR-4"),
    ("Infantry Fighting Vehicles", "BMP-1"),
    ("Armoured Personnel Carriers", "BTR-60"),
    ("Mine-Resistant Ambush Protected (MRAP) Vehicles", "Cougar"),
    ("Command Posts And Communications Stations", "R-142"),
    ("Self-Propelled Anti-Tank Missile Systems", "9P148"),
    ("Artillery Support Vehicles And Equipment", "1V13"),
    ("Self-Propelled Artillery", "2S1 Gvozdika"),
    ("Multiple Rocket Launchers", "BM-21 Grad"),
    ("Anti-Aircraft Guns", "ZU-23-2"),
    ("Self-Propelled Anti-Aircraft Guns", "ZSU-23-4 Shilka"),
    ("Surface-To-Air Missile Systems", "9K33 Osa"),
    ("Radars", "1L13"),
    ("Aircraft", "MiG-29"),
    ("Helicopters", "Mi-8"),
    ("Unmanned Combat Aerial Vehicles", "Bayraktar TB2"),
    ("Naval Ships and Submarines", "Haiduk"),
    ("Trucks, Vehicles, and Jeeps", "GAZ"),
]


def extract_status(status_part):
    counts = defaultdict(int)
    parts = re.split(r'[,\s]+and[,\s]+|[,\s]+', status_part)
    for p in parts:
        for key in ('destroyed', 'captured', 'damaged', 'abandoned'):
            if key in p:
                n = re.findall(r'\d+', p)
                counts[key.capitalize()] += int(n[0]) if n else 1
    if counts['Damaged'] > 0 and counts['Abandoned'] > 0:
        if counts['Damaged'] >= counts['Abandoned']:
            counts['Abandoned'] = 0
        else:
            counts['Damaged'] = 0
    return counts


def scrape() -> pd.DataFrame:
    log.info(f"Scraping (UKR): {URL}")
    try:
        # Kozos lekero ujraprobalkozassal — lasd fetch_with_backoff.py.
        # A 2026-08-04-i futas azert szallt el, mert egyetlen keres ment,
        # a Google botvedelme pedig 429-cel valaszolt.
        resp = fetch(URL)
        if resp is None:
            log.error("Nem sikerult lekerni az oldalt tobb kiserlet utan sem.")
            return None
    except Exception as e:
        log.error(f"Hálózati hiba: {e}")
        return None

    soup = BeautifulSoup(resp.content, 'html.parser')
    main = soup.find('div', class_='post-body entry-content')
    if not main:
        log.error("Nem található post-body. Változott az Oryx oldal struktúrája?")
        return None

    data = defaultdict(lambda: {'Destroyed': 0, 'Captured': 0, 'Damaged': 0,
                                 'Abandoned': 0, 'Sum_Loss': 0})
    for title in main.find_all('h3'):
        if not title.text.strip():
            continue
        sib = title.find_next_sibling()
        while sib and sib.name != 'h3':
            if sib.name in ['p', 'div', 'ul', 'ol']:
                items = sib.get_text(separator='\n', strip=True).split('\n')
                cur_type = None
                for item in items:
                    item = item.strip()
                    if ':' in item:
                        cur_type = re.sub(r'^\d+\s+', '', item.split(':')[0].strip())
                    elif '(' in item and cur_type:
                        sp = item.split('(')[-1].split(')')[0].strip()
                        for status, count in extract_status(sp).items():
                            data[cur_type][status] += count
                            data[cur_type]['Sum_Loss'] += count
            sib = sib.find_next_sibling()

    df = pd.DataFrame.from_dict(data, orient='index').reset_index()
    df.rename(columns={'index': 'Type'}, inplace=True)

    for cat_name, first_type in CATEGORIES:
        mask = df['Type'].str.startswith(first_type, na=False)
        if mask.any():
            idx = df[mask].index[0]
            cat_row = pd.DataFrame([{'Type': cat_name, 'Destroyed': None, 'Captured': None,
                                      'Damaged': None, 'Abandoned': None, 'Sum_Loss': None}])
            df = pd.concat([df.iloc[:idx], cat_row, df.iloc[idx:]]).reset_index(drop=True)

    log.info(f"Scraping kész: {len(df)} sor")
    return df


def save_to_excel(df: pd.DataFrame, filename: str):
    current_date = datetime.now().strftime('%Y-%m-%d')
    os.makedirs(os.path.dirname(os.path.abspath(filename)), exist_ok=True)

    if os.path.exists(filename):
        wb = load_workbook(filename)
        if current_date in wb.sheetnames:
            log.info(f"Már létezik lap erre a dátumra: {current_date} — kihagyva.")
            return
    else:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    ws = wb.create_sheet(current_date)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    wb.save(filename)
    log.info(f"Mentve: {current_date} → {filename}")


def run(filename: str = None):
    if not filename:
        filename = DEFAULT_FILE
    df = scrape()
    if df is not None:
        save_to_excel(df, filename)
    return df is not None


if __name__ == '__main__':
    ok = run(sys.argv[1] if len(sys.argv) > 1 else None)
    sys.exit(0 if ok else 1)
