import requests
from bs4 import BeautifulSoup
from collections import defaultdict
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

# URL of the webpage to scrape
url = "https://www.oryxspioenkop.com/2022/02/attack-on-europe-documenting-equipment.html"

# Send a GET request to the URL
response = requests.get(url)


def extract_status(status_part):
    """Extract status and count from status_part"""
    status_counts = defaultdict(int)
    status_split = re.split(r'[,\s]+and[,\s]+|[,\s]+', status_part)

    for part in status_split:
        if 'destroyed' in part:
            count = re.findall(r'\d+', part)
            count = int(count[0]) if count else 1
            status_counts['Destroyed'] += count
        elif 'captured' in part:
            count = re.findall(r'\d+', part)
            count = int(count[0]) if count else 1
            status_counts['Captured'] += count
        elif 'damaged' in part:
            count = re.findall(r'\d+', part)
            count = int(count[0]) if count else 1
            status_counts['Damaged'] += count
        elif 'abandoned' in part:
            count = re.findall(r'\d+', part)
            count = int(count[0]) if count else 1
            status_counts['Abandoned'] += count

    # Ensure no double counting
    if status_counts['Damaged'] > 0 and status_counts['Abandoned'] > 0:
        if status_counts['Damaged'] >= status_counts['Abandoned']:
            status_counts['Abandoned'] = 0
        else:
            status_counts['Damaged'] = 0

    return status_counts


def save_to_excel(df, filename):
    # Get current date and week number
    current_date = datetime.now().strftime("%Y-%m-%d")
    current_week = datetime.now().isocalendar()[1]

    if os.path.exists(filename):
        wb = load_workbook(filename)
        if str(current_week) in wb.sheetnames:
            print(f"Data for week {current_week} already exists. No changes made.")
            return
    else:
        wb = Workbook()

    # Create a new sheet with the current date as name
    ws = wb.create_sheet(current_date)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Save the workbook
    wb.save(filename)


def update_summation_sheet(filename, df):
    wb = load_workbook(filename)
    if 'Summation' not in wb.sheetnames:
        ws_summary = wb.create_sheet('Summation')
        ws_summary.append(['Type', 'Destroyed', 'Captured', 'Damaged', 'Abandoned', 'Sum_Loss'])
    else:
        ws_summary = wb['Summation']

    # Get the last week's data from the previous week
    prev_week = datetime.now().isocalendar()[1] - 1
    if str(prev_week) in wb.sheetnames:
        df_prev = pd.read_excel(filename, sheet_name=str(prev_week))
        df_prev.set_index('Type', inplace=True)
        df.set_index('Type', inplace=True)
        df_diff = df - df_prev
        df_diff.reset_index(inplace=True)
        df_diff['Week'] = prev_week
        df_diff['Date'] = datetime.now().strftime("%Y-%m-%d")
    else:
        df_diff = df.copy()
        df_diff['Week'] = datetime.now().isocalendar()[1]
        df_diff['Date'] = datetime.now().strftime("%Y-%m-%d")

    for r in dataframe_to_rows(df_diff, index=False, header=False):
        ws_summary.append(r)

    # Save the workbook
    wb.save(filename)


# Check if the request was successful
if response.status_code == 200:
    # Parse the HTML content of the webpage
    soup = BeautifulSoup(response.content, "html.parser")

    # Find the main container that holds the data (adjust based on actual structure)
    main_content = soup.find('div', class_='post-body entry-content')

    if main_content:
        # Find all the headings (h3 tags) which might be the titles
        titles = main_content.find_all('h3')

        equipment_data = defaultdict(
            lambda: {'Destroyed': 0, 'Captured': 0, 'Damaged': 0, 'Abandoned': 0, 'Sum_Loss': 0})

        # Iterate over each title and process the associated data
        for title in titles:
            title_text = title.text.strip()

            # Skip empty titles
            if not title_text:
                continue

            # Find the next sibling elements until the next h3 (or end of content)
            sibling = title.find_next_sibling()
            while sibling and sibling.name != 'h3':
                if sibling.name in ['p', 'div', 'ul', 'ol']:
                    items = sibling.get_text(separator="\n", strip=True).split("\n")
                    current_type = None

                    for item in items:
                        item = item.strip()
                        if ':' in item:
                            # New equipment type
                            current_type = re.sub(r'^\d+\s+', '', item.split(':')[0].strip())
                        elif '(' in item and current_type:
                            # Loss information
                            status_part = item.split('(')[-1].split(')')[0].strip()
                            status_counts = extract_status(status_part)

                            for status, count in status_counts.items():
                                if status in equipment_data[current_type]:
                                    equipment_data[current_type][status] += count
                                    equipment_data[current_type]['Sum_Loss'] += count
                sibling = sibling.find_next_sibling()

        # Create a DataFrame from the collected data
        df = pd.DataFrame.from_dict(equipment_data, orient='index').reset_index()
        df.rename(columns={'index': 'Type'}, inplace=True)

        # Categories to insert and their respective positions
        categories = [
            ("Tanks", "T-54"),
            ("Armoured Fighting Vehicles", "BMPT Terminator"),
            ("Infantry Fighting Vehicles", "BMP-1(P)"),
            ("Armoured Personnel Carriers", "BTR-50"),
            ("Mine-Resistant Ambush Protected (MRAP) Vehicles", "KamAZ-63968 Typhoon"),
            ("Infantry Mobility Vehicles", "BPM-97 Vystrel"),
            ("Command Posts And Communications Stations", "BMP-1KSh"),
            ("Engineering Vehicles And Equipment", "UR-67 demining charge"),
            ("Unmanned Ground Vehicles", "Uran-6 demining robot system"),
            ("Self-Propelled Anti-Tank Missile Systems", "9P148 Konkurs"),
            ("Artillery Support Vehicles And Equipment", "1V110 BM-21 Grad battery command vehicle"),
            ("Towed Artillery", "82mm 2B9 Vasilek automatic cannon"),
            ("Self-Propelled Artillery", "120mm 2S9 Nona"),
            ("Multiple Rocket Launchers", "122mm BM-21 Grad"),
            ("Anti-Aircraft Guns", "23mm ZU-23-2"),
            ("Self-Propelled Anti-Aircraft Guns", "BTR-ZD Skrezhet"),
            ("Surface-To-Air Missile Systems", "9K33 Osa"),
            ("Radars", "9S36"),
            ("Jammers And Deception Systems", "R-325BMV jamming station"),
            ("Aircraft", "MiG-31BM fighter aircraft"),
            ("Helicopters", "Mi-8 transport helicopter"),
            ("Unmanned Combat Aerial Vehicles", "Orion"),
            ("Reconnaissance Unmanned Aerial Vehicles", "forpost"),
            ("Naval Ships and Submarines", "Project 1164 Slava-class guided missile cruiser"),
            ("Trucks, Vehicles, and Jeeps", "GAZ-51"),
        ]

        # Insert categories as new rows at the specified positions
        for category, before_type in categories:
            # Escape parentheses in the before_type string to avoid regex groups
            before_type_escaped = re.escape(before_type)
            idx = df.index[df['Type'].str.contains(before_type_escaped, na=False)].min()
            if pd.notna(idx):
                new_row = pd.DataFrame([{'Type': category, 'Destroyed': '', 'Captured': '', 'Damaged': '',
                                         'Abandoned': '', 'Sum_Loss': ''}])
                df = pd.concat([df.iloc[:idx], new_row, df.iloc[idx:]]).reset_index(drop=True)

        # Save data to the Excel file
        filename = 'equipment_losses_with_categories.xlsx'
        save_to_excel(df, filename)

        # Update the summation sheet with the differences
        update_summation_sheet(filename, df)

        print("Data has been saved to 'equipment_losses_with_categories.xlsx'")
else:
    print(f"Failed to retrieve the webpage. Status code: {response.status_code}")