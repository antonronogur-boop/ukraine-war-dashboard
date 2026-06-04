"""
analysis_losses.py  –  Heti delta kalkulátor (RUS és UKR egyaránt)

Használat:
    python analysis_losses.py data/xyz_RUS.xlsx
    python analysis_losses.py data/xyz_UKR.xlsx

Mit csinál:
    - Minden dátum-nevű lapból kiszámolja a hét/hét változást
    - Eredményt a 'summary' lapra írja az xlsx-be
    - Futás után a 'summary' lap tartalmazza az összes heti deltát
"""

import sys
import re
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)


def run(filename: str):
    log.info(f"Feldolgozás: {filename}")
    wb = load_workbook(filename)
    sheetnames = wb.sheetnames

    # Csak dátum-nevű lapok (YYYY-MM-DD)
    date_sheets = sorted([s for s in sheetnames if re.match(r'\d{4}-\d{2}-\d{2}', s)])
    if not date_sheets:
        log.error("Nem találhatók dátum-nevű lapok.")
        return

    log.info(f"  {len(date_sheets)} heti lap találva: {date_sheets[0]} → {date_sheets[-1]}")

    aggregate_rows = []

    for i, sheetname in enumerate(date_sheets):
        current_df = pd.read_excel(filename, sheet_name=sheetname)

        if i == 0:
            base = current_df.copy()
            base['Week'] = 0
            base['Date'] = sheetname
            aggregate_rows.append(base)
            log.info(f"  Alap hét: {sheetname} ({len(base)} sor)")
        else:
            prev_df = pd.read_excel(filename, sheet_name=date_sheets[i - 1])
            cur = current_df.set_index('Type')
            prv = prev_df.set_index('Type')

            # Csak a számoszlopok különbsége
            numeric_cols = [c for c in cur.columns if c != 'Type']
            diff = cur[numeric_cols].subtract(prv[numeric_cols], fill_value=0)

            # Csak a sorokat tartjuk ahol van változás
            changed = diff[(diff != 0).any(axis=1)].copy()

            if not changed.empty:
                changed = changed.reset_index()
                changed['Week'] = i
                changed['Date'] = sheetname
                aggregate_rows.append(changed)
                log.info(f"  {sheetname}: {len(changed)} változó sor")
            else:
                log.info(f"  {sheetname}: nincs változás")

    if not aggregate_rows:
        log.warning("Nincs feldolgozható adat.")
        return

    aggregate_df = pd.concat(aggregate_rows, ignore_index=True)

    # Summary lap frissítése
    if 'summary' in wb.sheetnames:
        del wb['summary']
    ws_summary = wb.create_sheet('summary')

    # Kivonat: alap + deltás elrendezés (Type x Week_N_Col formátum)
    base_df = aggregate_df[aggregate_df['Week'] == 0][['Type', 'Destroyed', 'Captured',
                                                        'Damaged', 'Abandoned', 'Sum_Loss']].copy()
    base_df['Week'] = 0
    base_df['Date'] = date_sheets[0]

    summary_df = base_df.copy()

    for i in range(1, len(date_sheets)):
        diff_df = aggregate_df[(aggregate_df['Week'] == i)].set_index('Type')
        for col in ['Destroyed', 'Captured', 'Damaged', 'Abandoned', 'Sum_Loss']:
            col_name = f'{col}_Week_{i}'
            summary_df[col_name] = summary_df['Type'].map(
                diff_df[col] if col in diff_df.columns else pd.Series(dtype=float)
            )

    for r in dataframe_to_rows(summary_df, index=False, header=True):
        ws_summary.append(r)

    wb.save(filename)
    log.info(f"Kész. Summary lap mentve: {filename}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Használat: python analysis_losses.py <fájlnév.xlsx>")
        sys.exit(1)
    run(sys.argv[1])
