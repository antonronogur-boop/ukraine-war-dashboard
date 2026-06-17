"""
scraper_RUS.py  –  Orosz Oryx adatok scraper (javított verzió)

Használat:
    python scraper_RUS.py                    # default: data/xyz_RUS.xlsx
    python scraper_RUS.py data/xyz_RUS.xlsx  # explicit elérési út

Javítások az eredeti w_websc_RUS.py-hoz képest:
    - Dátum alapú duplikáció-ellenőrzés (az eredeti hétsszámot nézett — bug)
    - Fájlnév parancssorból vagy default
    - Hibakezelés hálózati hiba esetén
    - Üres Sheet törlése új fájlnál
    - Kategória beillesztés megbízhatóbb
"""

import sys, os, re, logging
import pandas as pd
import requests
from bs4 import BeautifulSoup
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)

URL = "https://www.oryxspioenkop.com/2022/02/attack-on-europe-documenting-equipment.html"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_FILE = os.path.normpath(os.path.join(SCRIPT_DIR, '..', 'data', 'xyz_RUS.xlsx'))

CATEGORIES = [
    ("Tanks", "T-54"),
    ("Armoured Fighting Vehicles", "BMPT Terminator"),
    ("Infantry Fighting Vehicles", "BMP-1(P)"),
    ("Armoured Personnel Carriers", "BTR-50"),
    ("Mine-Resistant Ambush Protected (MRAP) Vehicles", "KamAZ-63968 Typhoon"),
    ("Command Posts And Communications Stations", "BMP-1KSh"),
    ("Self-Propelled Anti-Tank Missile Systems", "9P148 Konkurs"),
    ("Artillery Support Vehicles And Equipment", "1V110 BM-21 Grad battery command vehicle"),
    ("Self-Propelled Artillery", "120mm 2S9 Nona"),
    ("Multiple Rocket Launchers", "122mm BM-21 Grad"),
    ("Anti-Aircraft Guns", "23mm ZU-23-2"),
    ("Self-Propelled Anti-Aircraft Guns", "BTR-ZD Skrezhet"),
    ("Surface-To-Air Missile Systems", "9K33 Osa"),
    ("Radars", "9S36"),
    ("Jammers And Deception Systems", "R-325BMV jamming station"),
    ("Aircraft", "MiG-31BM fighter aircraft"),
    ("Helicopters", "Mi-8 transport helicopter"),
    ("Unmanned Combat Aerial Vehicles", "Orion"),
    ("Naval Ships and Submarines", "Project 1164 Slava-class guided missile cruiser"),
    ("Trucks, Vehicles, and Jeeps", "GAZ-51"),
]


def extract_status(status_part):
    counts = defaultdict(int)
    parts = re.split(r'[,\s]+and[,\s]+|[,\s]+', status_part)
    for p in parts:
        for key in ('destroyed', 'captured', 'damaged', 'abandoned'):
            if key in p:
                n = re.findall(r'\d+', p)
                counts[key.capitalize()] += int(n[0]) if n else 1
    if counts['Damaged'] > 0 and counts['Abandoned'] > 0:
        if counts['Damaged'] >= counts['Abandoned']:
            counts['Abandoned'] = 0
        else:
            counts['Damaged'] = 0
    return counts


def scrape() -> pd.DataFrame:
    log.info(f"Scraping (RUS): {URL}")
    try:
        resp = requests.get(URL, timeout=30, headers={'User-Agent': 'Mozilla/5.0'})
        resp.raise_for_status()
    except Exception as e:
        log.error(f"Hálózati hiba: {e}")
        return None

    soup = BeautifulSoup(resp.content, 'html.parser')
    main = soup.find('div', class_='post-body entry-content')
    if not main:
        log.error("Nem található post-body. Változott az Oryx oldal struktúrája?")
        return None

    data = defaultdict(lambda: {'Destroyed': 0, 'Captured': 0, 'Damaged': 0,
                                 'Abandoned': 0, 'Sum_Loss': 0})
    for title in main.find_all('h3'):
        if not title.text.strip():
            continue
        sib = title.find_next_sibling()
        while sib and sib.name != 'h3':
            if sib.name in ['p', 'div', 'ul', 'ol']:
                items = sib.get_text(separator='\n', strip=True).split('\n')
                cur_type = None
                for item in items:
                    item = item.strip()
                    if ':' in item:
                        cur_type = re.sub(r'^\d+\s+', '', item.split(':')[0].strip())
                    elif '(' in item and cur_type:
                        sp = item.split('(')[-1].split(')')[0].strip()
                        for status, count in extract_status(sp).items():
                            data[cur_type][status] += count
                            data[cur_type]['Sum_Loss'] += count
            sib = sib.find_next_sibling()

    df = pd.DataFrame.from_dict(data, orient='index').reset_index()
    df.rename(columns={'index': 'Type'}, inplace=True)

    for cat_name, first_type in CATEGORIES:
        mask = df['Type'].str.startswith(first_type, na=False)
        if mask.any():
            idx = df[mask].index[0]
            cat_row = pd.DataFrame([{'Type': cat_name, 'Destroyed': None, 'Captured': None,
                                      'Damaged': None, 'Abandoned': None, 'Sum_Loss': None}])
            df = pd.concat([df.iloc[:idx], cat_row, df.iloc[idx:]]).reset_index(drop=True)

    log.info(f"Scraping kész: {len(df)} sor")
    return df


def save_to_excel(df: pd.DataFrame, filename: str):
    current_date = datetime.now().strftime('%Y-%m-%d')
    os.makedirs(os.path.dirname(os.path.abspath(filename)), exist_ok=True)

    if os.path.exists(filename):
        wb = load_workbook(filename)
        if current_date in wb.sheetnames:
            log.info(f"Már létezik lap erre a dátumra: {current_date} — kihagyva.")
            return
    else:
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

    ws = wb.create_sheet(current_date)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    wb.save(filename)
    log.info(f"Mentve: {current_date} → {filename}")


def run(filename: str = None):
    if not filename:
        filename = DEFAULT_FILE
    df = scrape()
    if df is not None:
        save_to_excel(df, filename)
    return df is not None


if __name__ == '__main__':
    ok = run(sys.argv[1] if len(sys.argv) > 1 else None)
    sys.exit(0 if ok else 1)
