"""
mod_excel_export.py  –  MoD napi adatok Excel adatbázisba mentése

Forrás:  Az index.html DAILY tömbje (ami a minfin.com.ua-ról gyűjtött adatokat tartalmazza)
Kimenet: data/mod_losses.xlsx  (minden futáskor egy új sorral bővül)

Futtatás:
    python scripts/mod_excel_export.py               # default
    python scripts/mod_excel_export.py index.html    # explicit HTML forrás

Logika:
    1. Beolvassa az index.html DAILY tömbjét (JSON parse)
    2. Beolvassa a meglévő mod_losses.xlsx-et (ha van)
    3. Hozzáfűzi az új sorokat (amelyek még nem szerepelnek)
    4. Elmenti

Táblázat oszlopai:
    date, war_day,
    personnel_total, personnel_daily,
    tanks_total, tanks_daily,
    afv_total, afv_daily,
    artillery_total, artillery_daily,
    mlrs_total, mlrs_daily,
    airdef_total, airdef_daily,
    planes_total, planes_daily,
    helicopters_total, helicopters_daily,
    uav_total, uav_daily,
    missiles_total, missiles_daily,
    ships_total, ships_daily,
    submarines_total, submarines_daily,
    vehicles_total, vehicles_daily,
    special_total, special_daily,
    robots_total, robots_daily
"""

import sys
import os
import re
import json
import logging
from datetime import date as Date, datetime

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
except ImportError:
    log.error('openpyxl nem elérhető. Telepítés: pip install openpyxl')
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_HTML  = os.path.normpath(os.path.join(SCRIPT_DIR, '..', 'index.html'))
DEFAULT_EXCEL = os.path.normpath(os.path.join(SCRIPT_DIR, '..', 'data', 'mod_losses.xlsx'))

WAR_START = Date(2022, 2, 24)

FIELDS = [
    'personnel', 'tanks', 'afv', 'artillery', 'mlrs', 'airdef',
    'planes', 'helicopters', 'uav', 'missiles', 'ships',
    'submarines', 'vehicles', 'special', 'robots'
]

HEADERS = ['date', 'war_day', 'gap_days'] + [
    col for f in FIELDS for col in (f + '_total', f + '_daily')
]


def war_day(date_str: str) -> int:
    d = Date.fromisoformat(date_str)
    return (d - WAR_START).days + 1


def extract_daily_from_html(html_path: str) -> list[dict]:
    """Kinyeri a DAILY tömböt az index.html-ből."""
    with open(html_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # A DAILY tömb keresése: const DAILY = [ ... ];
    m = re.search(r'const DAILY\s*=\s*(\[.*?\]);', content, re.DOTALL)
    if not m:
        log.error('Nem található a DAILY tömb az index.html-ben.')
        return []

    # JS objektumokat JSON-ná alakítjuk (idézőjelek hozzáadása a kulcsokhoz)
    js_array = m.group(1)
    json_str = re.sub(r'(\w+):', r'"\1":', js_array)
    # Dupla idézőjelek javítása (ha a kulcs már idézőjeles volt)
    json_str = re.sub(r'"([^"]+)":', r'"\1":', json_str)

    try:
        data = json.loads(json_str)
    except json.JSONDecodeError as e:
        log.error(f'JSON parse hiba: {e}')
        return []

    log.info(f'{len(data)} rekord kinyerve az index.html DAILY tömbjéből')
    return data


def compute_rows(daily_data: list[dict]) -> list[list]:
    """
    Kiszámolja a napi deltákat és összerakja a sorokat.

    Adatrés-normalizálás: ha a scraper egy vagy több napot kihagyott
    (pl. máj. 31 → jún. 6, 6 napos rés), a két bejegyzés közti nyers
    különbség a KIHAGYOTT NAPOK ÖSSZESÍTETT vesztesége, nem egyetlen
    napé. Ha ezt egy az egyben "napi" értékként mentenénk, mesterséges
    tüskét hoznánk létre az adatbázisban. Ehelyett egyenletesen szétosztjuk
    a rés napjaira (nyers delta ÷ rés hossza), és a gap_days oszlopban
    jelöljük, hogy az adott sor becsült napi átlagot tartalmaz.
    """
    rows = []
    for i, d in enumerate(daily_data):
        prev = daily_data[i - 1] if i > 0 else None
        if prev:
            gap = max(1, (Date.fromisoformat(d['date']) - Date.fromisoformat(prev['date'])).days)
        else:
            gap = 1
        row = [d['date'], war_day(d['date']), gap]
        for f in FIELDS:
            total = d.get(f, 0) or 0
            raw = max(0, total - (prev.get(f, 0) or 0)) if prev else 0
            # Robots spike javítása: ha az előző 0 volt és az ugrás >100, az nem valódi veszteség
            if f == 'robots' and (prev is None or (prev.get(f, 0) or 0) == 0) and raw > 100:
                raw = 0
            # Több napos rés esetén egyenletes szétosztás napi átlagként
            daily = round(raw / gap) if gap > 1 else raw
            row.extend([total, daily])
        rows.append(row)
    return rows


def load_existing_dates(xlsx_path: str) -> set:
    """Visszaadja a már mentett dátumokat."""
    if not os.path.exists(xlsx_path):
        return set()
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    dates = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            dates.add(str(row[0]))
    wb.close()
    return dates


def save_to_excel(rows: list[list], xlsx_path: str):
    """Hozzáfűzi az új sorokat az Excel fájlhoz."""
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

    existing_dates = load_existing_dates(xlsx_path)

    new_rows = [r for r in rows if r[0] not in existing_dates]
    if not new_rows:
        log.info('Nincs új adat — az Excel már naprakész.')
        return 0

    if os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path)
        ws = wb.active
        # Migráció: ha a régi fájlban nincs még 'gap_days' oszlop, beszúrjuk
        # a 3. helyre (régi sorok esetén 1-es alapértékkel — ezekről nem
        # tudjuk utólag, hogy volt-e rés, de a korábbi automatikus futások
        # napi rendszerességgel mentek, tehát az 1 a legjobb feltételezés).
        existing_header = [c.value for c in ws[1]]
        if existing_header and 'gap_days' not in existing_header:
            ws.insert_cols(3)
            ws.cell(row=1, column=3, value='gap_days')
            ws.cell(row=1, column=3).font = openpyxl.styles.Font(bold=True)
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value:
                    ws.cell(row=r, column=3, value=1)
            log.info("Migráció: 'gap_days' oszlop hozzáadva a meglévő adatbázishoz (régi sorok = 1).")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'MoD veszteségek'
        ws.append(HEADERS)
        # Fejléc formázása
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

    for row in new_rows:
        ws.append(row)

    # Dátum oszlop formázása
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for cell in row:
            cell.number_format = 'YYYY-MM-DD'

    wb.save(xlsx_path)
    log.info(f'{len(new_rows)} új sor hozzáfűzve → {xlsx_path}')
    return len(new_rows)


def run(html_path: str = None, excel_path: str = None) -> int:
    if not html_path:
        html_path = DEFAULT_HTML
    if not excel_path:
        excel_path = DEFAULT_EXCEL

    log.info(f'HTML forrás: {html_path}')
    log.info(f'Excel cél:   {excel_path}')

    daily = extract_daily_from_html(html_path)
    if not daily:
        return 0

    rows = compute_rows(daily)
    added = save_to_excel(rows, excel_path)
    return added


if __name__ == '__main__':
    html  = sys.argv[1] if len(sys.argv) > 1 else None
    excel = sys.argv[2] if len(sys.argv) > 2 else None
    added = run(html, excel)
    log.info(f'Kész. {added} új sor mentve.')
    sys.exit(0)
