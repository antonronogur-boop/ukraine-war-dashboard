"""
collector.py  –  Összesítő lap generátor (RUS és UKR egyaránt)

Használat:
    python collector.py data/xyz_RUS.xlsx
    python collector.py data/xyz_UKR.xlsx

Mit csinál:
    - A 'summary' lapból kinyeri a Sum_Loss_Week_N oszlopokat
    - Csak azokat a sorokat tartja meg ahol volt legalább 1 pozitív veszteség
    - Eredményt a 'C_SUM' lapra írja
    - NINCS hardcoded típuslista — minden Oryx-on szereplő típust automatikusan követ
"""

import sys
import logging
import openpyxl

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)


def run(filename: str):
    log.info(f"Collector futtatása: {filename}")
    wb = openpyxl.load_workbook(filename)

    if 'summary' not in wb.sheetnames:
        log.error("Nincs 'summary' lap. Futtasd előbb az analysis_losses.py-t.")
        return

    summary = wb['summary']

    # C_SUM lap újragenerálása
    if 'C_SUM' in wb.sheetnames:
        del wb['C_SUM']
    c_sum = wb.create_sheet('C_SUM')

    # Sum_Loss_Week_N oszlopok megkeresése
    header = [cell.value for cell in summary[1]]
    loss_col_indices = [i for i, h in enumerate(header)
                        if h and 'Sum_Loss_Week_' in str(h)]

    if not loss_col_indices:
        log.error("Nem találhatók Sum_Loss_Week_N oszlopok a summary lapon.")
        return

    log.info(f"  {len(loss_col_indices)} heti oszlop találva")

    # Fejléc írása
    c_sum.append(['Type'] + [header[i] for i in loss_col_indices])

    kept = 0
    skipped = 0

    for row in summary.iter_rows(min_row=2, max_row=summary.max_row, values_only=True):
        if not row[0]:
            continue

        type_name = str(row[0]).strip()
        losses = [row[i] for i in loss_col_indices]

        # Kategória fejléc sorok kihagyása (pl. "Tanks", "Aircraft")
        # Ezeket felismerjük: csak az első oszlop van kitöltve, a többi None
        if all(v is None for v in losses):
            skipped += 1
            continue

        # Csak pozitív veszteséggel rendelkező típusok
        has_loss = any(v is not None and isinstance(v, (int, float)) and v > 0
                       for v in losses)
        if has_loss:
            c_sum.append([type_name] + losses)
            kept += 1

    wb.save(filename)
    log.info(f"C_SUM kész: {kept} típus megtartva, {skipped} kihagyva — {filename}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Használat: python collector.py <fájlnév.xlsx>")
        sys.exit(1)
    run(sys.argv[1])
