"""
oryx_export.py
Exportálja az Oryx heti adatokat JSON formátumba a dashboard számára.
Futtatás: python oryx_export.py
Kimenet: oryx_data.json (ugyanabba a mappába)
"""

import openpyxl
import json
import re
from datetime import datetime

# ── Kategória mapping: Oryx neve → dashboard kategória ──────────────
CATEGORY_MAP = {
    'Tanks': 'tanks',
    'Armoured Fighting Vehicles': 'afv',
    'Infantry Fighting Vehicles': 'afv',
    'Armoured Personnel Carriers': 'afv',
    'Mine-Resistant Ambush Protected (MRAP) Vehicles': 'afv',
    'Self-Propelled Artillery': 'artillery',
    'Multiple Rocket Launchers': 'mlrs',
    'Anti-Aircraft Guns': 'airdef',
    'Self-Propelled Anti-Aircraft Guns': 'airdef',
    'Surface-To-Air Missile Systems': 'airdef',
    'Aircraft': 'planes',
    'Helicopters': 'helicopters',
    'Unmanned Combat Aerial Vehicles': 'uav',
    'Naval Ships and Submarines': 'naval',
    'Trucks, Vehicles, and Jeeps': 'vehicles',
    'Command Posts And Communications Stations': 'electronics',
    'Radars': 'electronics',
    'Jammers And Deception Systems': 'electronics',
    'Self-Propelled Anti-Tank Missile Systems': 'antitank',
    'Artillery Support Vehicles And Equipment': 'support',
}

DASHBOARD_CATS = ['tanks', 'afv', 'artillery', 'mlrs', 'airdef',
                  'planes', 'helicopters', 'uav', 'naval',
                  'vehicles', 'electronics', 'antitank', 'support']

CAT_LABELS = {
    'tanks': 'Harckocsi',
    'afv': 'Páncélozott harcjármű',
    'artillery': 'Tüzérség',
    'mlrs': 'MLRS',
    'airdef': 'Légvédelem',
    'planes': 'Repülő',
    'helicopters': 'Helikopter',
    'uav': 'UAV drón',
    'naval': 'Hajó / tengeralattjáró',
    'vehicles': 'Jármű / teherautó',
    'electronics': 'Elektr. / parancsnoki',
    'antitank': 'Páncéltörő rendszer',
    'support': 'Támogató eszköz',
}


def parse_file(path):
    """
    Beolvassa az xyz_*.xlsx fájlt.
    Visszaad: (week_dates, weekly_data)
      week_dates: ['2024-07-02', '2024-07-09', ...]
      weekly_data: [{cat: {D, C, Dam, A, total}, ...}, ...]  # egy dict/hét
    """
    wb = openpyxl.load_workbook(path, read_only=True)
    date_sheets = sorted([s for s in wb.sheetnames if re.match(r'\d{4}-\d{2}-\d{2}', s)])

    week_dates = []
    weekly_data = []

    for sheet_name in date_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        week_dates.append(sheet_name)
        week_totals = {cat: {'destroyed': 0, 'captured': 0, 'damaged': 0, 'abandoned': 0, 'total': 0}
                       for cat in DASHBOARD_CATS}

        # Az első sor a fejléc: Type, Destroyed, Captured, Damaged, Abandoned, Sum_Loss
        current_cat = None
        for row in rows[1:]:  # fejléc kihagyása
            if not row[0]:
                continue
            type_name = str(row[0]).strip()

            # Kategória fejléc sor (ha nincs számadat mellette)
            if row[1] is None and row[2] is None:
                current_cat = CATEGORY_MAP.get(type_name)
                continue

            if current_cat is None:
                continue

            # Típus adat sor
            d = int(row[1] or 0)
            c = int(row[2] or 0)
            dam = int(row[3] or 0)
            a = int(row[4] or 0)
            total = int(row[5] or 0)

            week_totals[current_cat]['destroyed'] += d
            week_totals[current_cat]['captured'] += c
            week_totals[current_cat]['damaged'] += dam
            week_totals[current_cat]['abandoned'] += a
            week_totals[current_cat]['total'] += total

        weekly_data.append(week_totals)

    wb.close()
    return week_dates, weekly_data


def build_output(dates, rus_data, ukr_data):
    """Összeállítja a JSON struktúrát."""

    # Heti összesítők kategóriánként (kumulatív)
    rus_weekly = []
    ukr_weekly = []

    for i, (date, rd, ud) in enumerate(zip(dates, rus_data, ukr_data)):
        rus_entry = {'date': date, 'week': i + 1}
        ukr_entry = {'date': date, 'week': i + 1}
        for cat in DASHBOARD_CATS:
            rus_entry[cat] = rd[cat]['total']
            ukr_entry[cat] = ud[cat]['total']
        rus_weekly.append(rus_entry)
        ukr_weekly.append(ukr_entry)

    # Kategória összesítők (utolsó hét)
    last_rus = rus_data[-1] if rus_data else {}
    last_ukr = ukr_data[-1] if ukr_data else {}

    rus_totals = {}
    ukr_totals = {}
    for cat in DASHBOARD_CATS:
        rus_totals[cat] = last_rus.get(cat, {})
        ukr_totals[cat] = last_ukr.get(cat, {})

    # Heti delta (új veszteségek hetente) — csak a total alapján
    rus_deltas = []
    ukr_deltas = []
    for i in range(1, len(rus_weekly)):
        rus_delta = {'date': dates[i], 'week': i + 1}
        ukr_delta = {'date': dates[i], 'week': i + 1}
        for cat in DASHBOARD_CATS:
            rus_delta[cat] = max(0, rus_weekly[i][cat] - rus_weekly[i-1][cat])
            ukr_delta[cat] = max(0, ukr_weekly[i][cat] - ukr_weekly[i-1][cat])
        rus_deltas.append(rus_delta)
        ukr_deltas.append(ukr_delta)

    return {
        'meta': {
            'source': 'Oryx (oryxspioenkop.com)',
            'source_type': 'Vizuálisan dokumentált, fotóval igazolt veszteségek',
            'note': 'Ez a lista csak azokat a veszteségeket tartalmazza, amelyeket képi bizonyítékkal igazoltak. A valós veszteségek magasabbak lehetnek.',
            'first_week': dates[0] if dates else None,
            'last_week': dates[-1] if dates else None,
            'total_weeks': len(dates),
            'categories': CAT_LABELS,
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
        },
        'russia': {
            'totals': rus_totals,
            'weekly_cumulative': rus_weekly,
            'weekly_delta': rus_deltas,
        },
        'ukraine': {
            'totals': ukr_totals,
            'weekly_cumulative': ukr_weekly,
            'weekly_delta': ukr_deltas,
        }
    }


def main():
    import os
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # GitHub Actions struktúra: data/ almappában vannak az xlsx fájlok
    data_dir = os.path.join(script_dir, 'data')
    if not os.path.exists(os.path.join(data_dir, 'xyz_RUS.xlsx')):
        data_dir = script_dir  # fallback: régi helyi elrendezés

    rus_path = os.path.join(data_dir, 'xyz_RUS.xlsx')
    ukr_path = os.path.join(data_dir, 'xyz_UKR.xlsx')
    out_path = os.path.join(script_dir, 'oryx_data.json')

    print('Beolvasom az orosz adatokat...')
    rus_dates, rus_data = parse_file(rus_path)
    print(f'  {len(rus_dates)} hét betöltve')

    print('Beolvasom az ukrán adatokat...')
    ukr_dates, ukr_data = parse_file(ukr_path)
    print(f'  {len(ukr_dates)} hét betöltve')

    # Ha eltér a két dátumsor, a közöset vesszük
    common = sorted(set(rus_dates) & set(ukr_dates))
    rus_idx = {d: i for i, d in enumerate(rus_dates)}
    ukr_idx = {d: i for i, d in enumerate(ukr_dates)}
    rus_common = [rus_data[rus_idx[d]] for d in common]
    ukr_common = [ukr_data[ukr_idx[d]] for d in common]

    print(f'Közös hetek: {len(common)}  ({common[0]} → {common[-1]})')

    output = build_output(common, rus_common, ukr_common)

    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    # Ellenőrzés: utolsó hét orosz tankjainak összege
    last = output['russia']['totals']
    print('\n── Utolsó hét orosz összesítő (Oryx) ──')
    for cat, label in CAT_LABELS.items():
        total = last.get(cat, {}).get('total', 0)
        if total > 0:
            d = last.get(cat, {}).get('destroyed', 0)
            c = last.get(cat, {}).get('captured', 0)
            print(f'  {label}: {total} (megsemmisített: {d}, zsákmányolt: {c})')

    print(f'\nMentve: {out_path}')
    print(f'Fájlméret: {os.path.getsize(out_path) // 1024} KB')


if __name__ == '__main__':
    main()
